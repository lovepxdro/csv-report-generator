from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analyzer import DatasetAnalysis
from src.dashboard import DashboardGenerator


class ExcelReportGenerator:
    def __init__(
        self,
        dataframe: pd.DataFrame,
        analysis: DatasetAnalysis,
        include_dashboard: bool = False,
    ):
        self.df = dataframe
        self.analysis = analysis
        self.include_dashboard = include_dashboard

    def generate(self, output_path: str) -> None:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(
            output,
            engine="openpyxl",
        ) as writer:
            self.df.to_excel(
                writer,
                index=False,
                sheet_name="Dados",
            )

            pd.DataFrame().to_excel(
                writer,
                index=False,
                sheet_name="Resumo",
            )

        workbook = load_workbook(output)

        data_sheet = workbook["Dados"]
        summary_sheet = workbook["Resumo"]

        self._format_data_sheet(data_sheet)
        self._build_summary_sheet(summary_sheet)

        if self.include_dashboard:
            dashboard_generator = DashboardGenerator(
                self.df,
                self.analysis,
            )

            dashboard_generator.build(workbook)

            workbook.active = workbook.sheetnames.index(
                "Dashboard"
            )
        else:
            workbook.active = workbook.sheetnames.index(
                "Resumo"
            )

        workbook.save(output)

    def _format_data_sheet(self, worksheet) -> None:
        self._format_header(worksheet)
        self._adjust_column_widths(worksheet)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column in self.analysis.datetime_columns:
            column_index = self.df.columns.get_loc(column) + 1

            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=row,
                    column=column_index,
                ).number_format = "DD/MM/YYYY"

    def _build_summary_sheet(self, worksheet) -> None:
        worksheet.sheet_view.showGridLines = False

        worksheet.merge_cells("A1:E2")

        title = worksheet["A1"]
        title.value = "Resumo do Dataset"

        title.font = Font(
            bold=True,
            size=20,
            color="FFFFFF",
        )

        title.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        title.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        total_nulls = sum(
            self.analysis.null_values.values()
        )

        total_cells = (
            self.analysis.rows
            * self.analysis.columns
        )

        null_percentage = (
            total_nulls / total_cells * 100
            if total_cells
            else 0
        )

        general_metrics = [
            ("Total de registros", self.analysis.rows),
            ("Total de colunas", self.analysis.columns),
            (
                "Colunas numéricas",
                len(self.analysis.numeric_columns),
            ),
            (
                "Colunas categóricas",
                len(self.analysis.categorical_columns),
            ),
            (
                "Colunas de data",
                len(self.analysis.datetime_columns),
            ),
            (
                "Valores ausentes",
                total_nulls,
            ),
            (
                "% de valores ausentes",
                null_percentage,
            ),
            (
                "Métrica principal",
                self.analysis.primary_metric or "N/A",
            ),
        ]

        row = 4

        for label, value in general_metrics:
            label_cell = worksheet.cell(
                row=row,
                column=1,
                value=label,
            )

            value_cell = worksheet.cell(
                row=row,
                column=2,
                value=value,
            )

            label_cell.font = Font(bold=True)

            if label == "% de valores ausentes":
                value_cell.number_format = "0.00%"

                value_cell.value = (
                    null_percentage / 100
                )

            row += 1

        row += 2

        if self.analysis.numeric_stats:
            row = self._write_section_title(
                worksheet,
                row,
                "Estatísticas Numéricas",
            )

            headers = [
                "Coluna",
                "Média",
                "Mínimo",
                "Máximo",
                "Soma",
            ]

            self._write_headers(
                worksheet,
                row,
                headers,
            )

            row += 1

            for column, stats in self.analysis.numeric_stats.items():
                worksheet.cell(
                    row=row,
                    column=1,
                    value=column,
                )

                values = [
                    stats["mean"],
                    stats["min"],
                    stats["max"],
                    stats["sum"],
                ]

                for index, value in enumerate(
                    values,
                    start=2,
                ):
                    cell = worksheet.cell(
                        row=row,
                        column=index,
                        value=value,
                    )

                    cell.number_format = '#,##0.00'

                row += 1

            row += 2

        if self.analysis.unique_values:
            row = self._write_section_title(
                worksheet,
                row,
                "Colunas Categóricas",
            )

            self._write_headers(
                worksheet,
                row,
                [
                    "Coluna",
                    "Valores únicos",
                ],
            )

            row += 1

            for column, unique_count in (
                self.analysis.unique_values.items()
            ):
                worksheet.cell(
                    row=row,
                    column=1,
                    value=column,
                )

                worksheet.cell(
                    row=row,
                    column=2,
                    value=unique_count,
                )

                row += 1

            row += 2

        if self.analysis.date_ranges:
            row = self._write_section_title(
                worksheet,
                row,
                "Períodos",
            )

            self._write_headers(
                worksheet,
                row,
                [
                    "Coluna",
                    "Data inicial",
                    "Data final",
                ],
            )

            row += 1

            for column, date_range in (
                self.analysis.date_ranges.items()
            ):
                worksheet.cell(
                    row=row,
                    column=1,
                    value=column,
                )

                start_cell = worksheet.cell(
                    row=row,
                    column=2,
                    value=date_range["min"],
                )

                end_cell = worksheet.cell(
                    row=row,
                    column=3,
                    value=date_range["max"],
                )

                start_cell.number_format = "DD/MM/YYYY"
                end_cell.number_format = "DD/MM/YYYY"

                row += 1

        self._adjust_column_widths(worksheet)

        worksheet.column_dimensions["A"].width = max(
            worksheet.column_dimensions["A"].width,
            24,
        )

    @staticmethod
    def _write_section_title(
        worksheet,
        row: int,
        title: str,
    ) -> int:
        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=5,
        )

        cell = worksheet.cell(
            row=row,
            column=1,
            value=title,
        )

        cell.font = Font(
            bold=True,
            size=13,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="4472C4",
        )

        return row + 1

    @staticmethod
    def _write_headers(
        worksheet,
        row: int,
        headers: list[str],
    ) -> None:
        for index, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=row,
                column=index,
                value=header,
            )

            cell.font = Font(bold=True)

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )

    @staticmethod
    def _format_header(worksheet) -> None:
        fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center"
            )

    @staticmethod
    def _adjust_column_widths(worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                if cell.value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                50,
            )
