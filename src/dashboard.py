import pandas as pd
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    Reference,
)
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)

from openpyxl.formatting.rule import ColorScaleRule

class DashboardGenerator:
    def __init__(
        self,
        dataframe: pd.DataFrame,
        analysis,
    ):
        self.df = dataframe
        self.analysis = analysis

    def build(self, workbook) -> None:
        dashboard = workbook.create_sheet(
            "Dashboard"
        )

        data_sheet = workbook.create_sheet(
            "DashboardData"
        )

        self._build_title(dashboard)

        if self.analysis.profile_type == "numeric":
            self._build_numeric_dashboard(
                dashboard,
                data_sheet,
            )

        elif self.analysis.profile_type == "categorical":
            self._build_categorical_dashboard(
                dashboard,
                data_sheet,
            )

        elif self.analysis.profile_type == "temporal":
            self._build_temporal_dashboard(
                dashboard,
                data_sheet,
            )

        else:
            self._build_mixed_dashboard(
                dashboard,
                data_sheet,
            )

        data_sheet.sheet_state = "hidden"
        dashboard.sheet_view.showGridLines = False

        for column in "ABCDEFGHIJKLMNOP":
            dashboard.column_dimensions[column].width = 14

    def _build_title(self, worksheet) -> None:
        worksheet.merge_cells("A1:O2")

        cell = worksheet["A1"]

        cell.value = (
            f"Dashboard — "
            f"{self.analysis.profile_type.title()}"
        )

        cell.font = Font(
            bold=True,
            size=20,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    def _build_numeric_dashboard(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        total_nulls = sum(
            self.analysis.null_values.values()
        )

        kpis = [
            (
                "Registros",
                self.analysis.rows,
            ),
            (
                "Variáveis numéricas",
                len(
                    self.analysis.numeric_columns
                ),
            ),
            (
                "Valores ausentes",
                total_nulls,
            ),
            (
                "Perfil",
                "Numérico",
            ),
        ]

        self._write_kpis(
            dashboard,
            kpis,
        )

        self._build_variability_chart(
            dashboard,
            data_sheet,
        )

        self._build_correlations_table(
            dashboard,
        )

        self._build_numeric_summary_table(
            dashboard,
        )
        
        self._build_correlation_matrix(
            dashboard,
        )

    def _build_mixed_dashboard(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        metric = self.analysis.primary_metric

        total_nulls = sum(
            self.analysis.null_values.values()
        )

        if metric:
            stats = (
                self.analysis.numeric_stats[
                    metric
                ]
            )

            kpis = [
                (
                    "Registros",
                    self.analysis.rows,
                ),
                (
                    f"Total {metric}",
                    stats["sum"],
                ),
                (
                    f"Média {metric}",
                    stats["mean"],
                ),
                (
                    "Valores ausentes",
                    total_nulls,
                ),
            ]

        else:
            kpis = [
                (
                    "Registros",
                    self.analysis.rows,
                ),
                (
                    "Colunas",
                    self.analysis.columns,
                ),
                (
                    "Numéricas",
                    len(
                        self.analysis.numeric_columns
                    ),
                ),
                (
                    "Valores ausentes",
                    total_nulls,
                ),
            ]

        self._write_kpis(
            dashboard,
            kpis,
        )

        current_row = 1

        current_row = (
            self._build_time_analysis(
                dashboard,
                data_sheet,
                current_row,
            )
        )

        current_row = (
            self._build_category_analysis(
                dashboard,
                data_sheet,
                current_row,
            )
        )

        self._build_category_distribution(
            dashboard,
            data_sheet,
            current_row,
        )

        self._build_top_categories(
            dashboard
        )

    def _write_kpis(
        self,
        worksheet,
        kpis,
    ) -> None:
        positions = [
            1,
            5,
            9,
            13,
        ]

        for (
            label,
            value,
        ), column in zip(
            kpis,
            positions,
        ):
            worksheet.merge_cells(
                start_row=4,
                start_column=column,
                end_row=4,
                end_column=column + 2,
            )

            worksheet.merge_cells(
                start_row=5,
                start_column=column,
                end_row=6,
                end_column=column + 2,
            )

            label_cell = worksheet.cell(
                row=4,
                column=column,
                value=label,
            )

            value_cell = worksheet.cell(
                row=5,
                column=column,
                value=value,
            )

            label_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="5B9BD5",
            )

            label_cell.font = Font(
                color="FFFFFF",
                bold=True,
            )

            label_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell.font = Font(
                bold=True,
                size=16,
            )

            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            if isinstance(value, pd.Timestamp):
                value_cell.number_format = "DD/MM/YYYY"

    def _build_variability_chart(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        stats = []

        for column, values in (
            self.analysis.numeric_stats.items()
        ):
            cv = values["cv"]

            if cv is None:
                continue

            stats.append(
                (
                    column,
                    cv,
                )
            )

        stats.sort(
            key=lambda item: item[1],
            reverse=True,
        )

        stats = stats[:10]

        if not stats:
            return

        data_sheet["A1"] = "Variável"
        data_sheet["B1"] = "Coeficiente de variação"

        for row, (column, cv) in enumerate(
            stats,
            start=2,
        ):
            data_sheet.cell(
                row=row,
                column=1,
                value=column,
            )

            data_sheet.cell(
                row=row,
                column=2,
                value=cv,
            ).number_format = "0.00"

        chart = BarChart()
        chart.type = "bar"

        chart.title = (
            "Maior variabilidade relativa"
        )

        chart.x_axis.title = (
            "Coeficiente de variação"
        )

        chart.y_axis.title = "Variável"

        chart.height = 10
        chart.width = 15

        data = Reference(
            data_sheet,
            min_col=2,
            min_row=1,
            max_row=len(stats) + 1,
        )

        categories = Reference(
            data_sheet,
            min_col=1,
            min_row=2,
            max_row=len(stats) + 1,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "A9",
        )

    def _build_correlations_table(
        self,
        worksheet,
    ) -> None:
        correlations = (
            self.analysis.top_correlations
        )

        if not correlations:
            return

        start_row = 9
        start_column = 10

        title = worksheet.cell(
            row=start_row,
            column=start_column,
            value="Top correlações",
        )

        title.font = Font(
            bold=True,
            size=14,
        )

        header_row = start_row + 2

        headers = [
            "Variável A",
            "Variável B",
            "Correlação",
        ]

        for index, header in enumerate(
            headers
        ):
            cell = worksheet.cell(
                row=header_row,
                column=start_column + index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="5B9BD5",
            )

        for index, item in enumerate(
            correlations,
            start=1,
        ):
            row = header_row + index

            worksheet.cell(
                row=row,
                column=start_column,
                value=item["column_a"],
            )

            worksheet.cell(
                row=row,
                column=start_column + 1,
                value=item["column_b"],
            )

            cell = worksheet.cell(
                row=row,
                column=start_column + 2,
                value=item["correlation"],
            )

            cell.number_format = "0.000"

    def _build_numeric_summary_table(
        self,
        worksheet,
    ) -> None:
        start_row = 25

        worksheet.cell(
            row=start_row,
            column=1,
            value="Resumo das variáveis",
        ).font = Font(
            bold=True,
            size=14,
        )

        headers = [
            "Variável",
            "Tipo",
            "Média",
            "Mínimo",
            "Máximo",
            "Desvio padrão",
            "CV",
        ]

        header_row = start_row + 2

        for index, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="5B9BD5",
            )

        for row_offset, (
            column,
            stats,
        ) in enumerate(
            self.analysis.numeric_stats.items(),
            start=1,
        ):
            row = header_row + row_offset

            worksheet.cell(
                row=row,
                column=1,
                value=column,
            )

            worksheet.cell(
                row=row,
                column=2,
                value=self.analysis.numeric_types[
                    column
                ],
            )

            values = [
                stats["mean"],
                stats["min"],
                stats["max"],
                stats["std"],
                stats["cv"],
            ]

            for column_index, value in enumerate(
                values,
                start=3,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column_index,
                    value=value,
                )

                if value is not None:
                    cell.number_format = "#,##0.00"

    def _build_time_analysis(
        self,
        dashboard,
        data_sheet,
        start_row,
    ):
        metric = self.analysis.primary_metric

        if (
            not self.analysis.datetime_columns
            or metric is None
        ):
            return start_row

        date_column = (
            self.analysis.datetime_columns[0]
        )

        grouped = (
            self.df[
                [date_column, metric]
            ]
            .dropna()
            .groupby(
                date_column,
                as_index=False,
            )[metric]
            .sum()
            .sort_values(
                date_column
            )
        )

        row = start_row

        data_sheet.cell(
            row=row,
            column=4,
            value=date_column,
        )

        data_sheet.cell(
            row=row,
            column=5,
            value=metric,
        )

        for _, record in grouped.iterrows():
            row += 1

            date_cell = data_sheet.cell(
                row=row,
                column=4,
                value=record[date_column],
            )

            date_cell.number_format = (
                "DD/MM/YYYY"
            )

            data_sheet.cell(
                row=row,
                column=5,
                value=float(
                    record[metric]
                ),
            )

        chart = LineChart()

        chart.title = (
            f"{metric.title()} "
            f"ao longo do tempo"
        )

        chart.height = 8
        chart.width = 14

        data = Reference(
            data_sheet,
            min_col=5,
            min_row=start_row,
            max_row=row,
        )

        categories = Reference(
            data_sheet,
            min_col=4,
            min_row=start_row + 1,
            max_row=row,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "A8",
        )

        return row + 3

    def _build_category_analysis(
        self,
        dashboard,
        data_sheet,
        start_row,
    ):
        metric = self.analysis.primary_metric
        category = (
            self._select_category_column()
        )

        if (
            metric is None
            or category is None
        ):
            return start_row

        grouped = (
            self.df[
                [category, metric]
            ]
            .dropna()
            .groupby(
                category,
                as_index=False,
            )[metric]
            .sum()
            .sort_values(
                metric,
                ascending=False,
            )
            .head(10)
        )

        row = start_row

        data_sheet.cell(
            row=row,
            column=7,
            value=category,
        )

        data_sheet.cell(
            row=row,
            column=8,
            value=metric,
        )

        for _, record in grouped.iterrows():
            row += 1

            data_sheet.cell(
                row=row,
                column=7,
                value=str(
                    record[category]
                ),
            )

            data_sheet.cell(
                row=row,
                column=8,
                value=float(
                    record[metric]
                ),
            )

        chart = BarChart()

        chart.title = (
            f"{metric.title()} "
            f"por {category.title()}"
        )

        chart.height = 8
        chart.width = 14

        data = Reference(
            data_sheet,
            min_col=8,
            min_row=start_row,
            max_row=row,
        )

        categories = Reference(
            data_sheet,
            min_col=7,
            min_row=start_row + 1,
            max_row=row,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "I8",
        )

        return row + 3

    def _build_category_distribution(
        self,
        dashboard,
        data_sheet,
        start_row,
    ):
        category = (
            self._select_category_column(
                max_unique=6
            )
        )

        if category is None:
            return

        counts = (
            self.df[category]
            .dropna()
            .value_counts()
        )

        row = start_row

        data_sheet.cell(
            row=row,
            column=10,
            value=category,
        )

        data_sheet.cell(
            row=row,
            column=11,
            value="Quantidade",
        )

        for value, count in counts.items():
            row += 1

            data_sheet.cell(
                row=row,
                column=10,
                value=str(value),
            )

            data_sheet.cell(
                row=row,
                column=11,
                value=int(count),
            )

        chart = PieChart()

        chart.title = (
            f"Distribuição por "
            f"{category.title()}"
        )

        chart.height = 8
        chart.width = 12

        data = Reference(
            data_sheet,
            min_col=11,
            min_row=start_row,
            max_row=row,
        )

        labels = Reference(
            data_sheet,
            min_col=10,
            min_row=start_row + 1,
            max_row=row,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(labels)

        dashboard.add_chart(
            chart,
            "A24",
        )

    def _build_top_categories(
        self,
        worksheet,
    ):
        metric = self.analysis.primary_metric
        category = (
            self._select_category_column()
        )

        if metric is None or category is None:
            return

        grouped = (
            self.df[
                [category, metric]
            ]
            .dropna()
            .groupby(
                category,
                as_index=False,
            )[metric]
            .sum()
            .sort_values(
                metric,
                ascending=False,
            )
            .head(5)
        )

        start_row = 24
        start_column = 9

        worksheet.cell(
            row=start_row,
            column=start_column,
            value=f"Top {category.title()}",
        ).font = Font(
            bold=True,
            size=14,
        )

        for index, header in enumerate(
            [
                category.title(),
                metric.title(),
            ]
        ):
            cell = worksheet.cell(
                row=start_row + 2,
                column=start_column + index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="5B9BD5",
            )

        for index, (
            _,
            record,
        ) in enumerate(
            grouped.iterrows(),
            start=1,
        ):
            row = (
                start_row
                + 2
                + index
            )

            worksheet.cell(
                row=row,
                column=start_column,
                value=str(
                    record[category]
                ),
            )

            cell = worksheet.cell(
                row=row,
                column=start_column + 1,
                value=float(
                    record[metric]
                ),
            )

            cell.number_format = "#,##0.00"

    def _select_category_column(
        self,
        max_unique=20,
    ):
        suitable = []

        for column in (
            self.analysis.categorical_columns
        ):
            unique_count = (
                self.analysis.unique_values[
                    column
                ]
            )

            if (
                2
                <= unique_count
                <= max_unique
            ):
                suitable.append(
                    (
                        column,
                        unique_count,
                    )
                )

        if not suitable:
            return None

        suitable.sort(
            key=lambda item: item[1]
        )

        return suitable[0][0]
        
    def _build_correlation_matrix(
        self,
        worksheet,
    ) -> None:
        numeric_columns = self.analysis.numeric_columns

        if len(numeric_columns) < 2:
            return

        correlation_matrix = (
            self.df[numeric_columns]
            .corr()
        )

        # Evita uma matriz gigantesca no dashboard.
        max_columns = 10
        selected_columns = (
            self._select_correlation_columns(
                limit=max_columns
            )
        )

        correlation_matrix = correlation_matrix.loc[
            selected_columns,
            selected_columns,
        ]

        start_row = 25
        start_column = 9

        worksheet.cell(
            row=start_row,
            column=start_column,
            value="Matriz de correlação",
        ).font = Font(
            bold=True,
            size=14,
        )

        header_row = start_row + 2

        for index, column in enumerate(
            selected_columns,
            start=1,
        ):
            worksheet.cell(
                row=header_row,
                column=start_column + index,
                value=column,
            ).font = Font(bold=True)

            worksheet.cell(
                row=header_row + index,
                column=start_column,
                value=column,
            ).font = Font(bold=True)

        for row_index, row_column in enumerate(
            selected_columns,
            start=1,
        ):
            for column_index, column_column in enumerate(
                selected_columns,
                start=1,
            ):
                value = correlation_matrix.loc[
                    row_column,
                    column_column,
                ]

                cell = worksheet.cell(
                    row=header_row + row_index,
                    column=start_column + column_index,
                    value=float(value),
                )

                cell.number_format = "0.00"

        first_data_row = header_row + 1
        last_data_row = header_row + len(selected_columns)

        first_data_column = start_column + 1
        last_data_column = start_column + len(selected_columns)

        start_cell = worksheet.cell(
            row=first_data_row,
            column=first_data_column,
        ).coordinate

        end_cell = worksheet.cell(
            row=last_data_row,
            column=last_data_column,
        ).coordinate

        worksheet.conditional_formatting.add(
            f"{start_cell}:{end_cell}",
            ColorScaleRule(
                start_type="num",
                start_value=-1,
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFEB84",
                end_type="num",
                end_value=1,
                end_color="63BE7B",
            ),
        )
        
    def _select_correlation_columns(
        self,
        limit: int = 10,
    ) -> list[str]:
        numeric_columns = (
            self.analysis.numeric_columns
        )

        if len(numeric_columns) <= limit:
            return numeric_columns

        correlation_matrix = (
            self.df[numeric_columns]
            .corr()
            .abs()
        )

        scores = {}

        for column in numeric_columns:
            correlations = (
                correlation_matrix[column]
                .drop(labels=[column])
                .dropna()
            )

            if correlations.empty:
                scores[column] = 0
            else:
                scores[column] = (
                    correlations.mean()
                )

        ranked = sorted(
            numeric_columns,
            key=lambda column: scores[column],
            reverse=True,
        )

        return ranked[:limit]
        
    def _build_categorical_dashboard(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        total_nulls = sum(
            self.analysis.null_values.values()
        )

        kpis = [
            (
                "Registros",
                self.analysis.rows,
            ),
            (
                "Variáveis categóricas",
                len(self.analysis.categorical_columns),
            ),
            (
                "Valores ausentes",
                total_nulls,
            ),
            (
                "Perfil",
                "Categórico",
            ),
        ]

        self._write_kpis(
            dashboard,
            kpis,
        )

        self._build_cardinality_chart(
            dashboard,
            data_sheet,
        )

        self._build_main_category_distribution(
            dashboard,
            data_sheet,
        )

        self._build_categorical_summary_table(
            dashboard,
        )
        
    def _build_cardinality_chart(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        values = sorted(
            self.analysis.unique_values.items(),
            key=lambda item: item[1],
            reverse=True,
        )[:10]

        if not values:
            return

        data_sheet["A1"] = "Coluna"
        data_sheet["B1"] = "Valores únicos"

        for row, (column, count) in enumerate(
            values,
            start=2,
        ):
            data_sheet.cell(
                row=row,
                column=1,
                value=column,
            )

            data_sheet.cell(
                row=row,
                column=2,
                value=count,
            )

        chart = BarChart()
        chart.type = "bar"

        chart.title = "Cardinalidade das colunas"
        chart.x_axis.title = "Valores únicos"
        chart.y_axis.title = "Coluna"

        chart.height = 9
        chart.width = 15

        data = Reference(
            data_sheet,
            min_col=2,
            min_row=1,
            max_row=len(values) + 1,
        )

        categories = Reference(
            data_sheet,
            min_col=1,
            min_row=2,
            max_row=len(values) + 1,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "A9",
        )
        
    def _select_main_categorical_column(
        self,
    ) -> str | None:
        candidates = []

        for column in (
            self.analysis.categorical_columns
        ):
            unique_count = (
                self.analysis.unique_values[
                    column
                ]
            )

            if 2 <= unique_count <= 15:
                candidates.append(
                    (
                        column,
                        unique_count,
                    )
                )

        if not candidates:
            return None

        candidates.sort(
            key=lambda item: item[1]
        )

        return candidates[0][0]
        
    def _build_main_category_distribution(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        column = (
            self._select_main_categorical_column()
        )

        if column is None:
            return

        counts = (
            self.df[column]
            .dropna()
            .value_counts()
            .head(10)
        )

        if counts.empty:
            return

        start_row = 1

        data_sheet.cell(
            row=start_row,
            column=4,
            value=column,
        )

        data_sheet.cell(
            row=start_row,
            column=5,
            value="Quantidade",
        )

        row = start_row

        for value, count in counts.items():
            row += 1

            data_sheet.cell(
                row=row,
                column=4,
                value=str(value),
            )

            data_sheet.cell(
                row=row,
                column=5,
                value=int(count),
            )

        chart = BarChart()
        chart.type = "col"

        chart.title = (
            f"Distribuição por {column.title()}"
        )

        chart.height = 9
        chart.width = 14

        data = Reference(
            data_sheet,
            min_col=5,
            min_row=start_row,
            max_row=row,
        )

        categories = Reference(
            data_sheet,
            min_col=4,
            min_row=start_row + 1,
            max_row=row,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "I9",
        )
        
    def _build_categorical_summary_table(
        self,
        worksheet,
    ) -> None:
        start_row = 27

        worksheet.cell(
            row=start_row,
            column=1,
            value="Resumo das variáveis categóricas",
        ).font = Font(
            bold=True,
            size=14,
        )

        headers = [
            "Coluna",
            "Valores únicos",
            "Mais frequente",
            "Frequência",
            "% do total",
            "Ausentes",
        ]

        header_row = start_row + 2

        for index, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="5B9BD5",
            )

        for offset, column in enumerate(
            self.analysis.categorical_columns,
            start=1,
        ):
            row = header_row + offset

            series = self.df[column]

            counts = (
                series
                .dropna()
                .value_counts()
            )

            most_frequent = (
                counts.index[0]
                if not counts.empty
                else None
            )

            frequency = (
                int(counts.iloc[0])
                if not counts.empty
                else 0
            )

            percentage = (
                frequency / self.analysis.rows
                if self.analysis.rows
                else 0
            )

            worksheet.cell(
                row=row,
                column=1,
                value=column,
            )

            worksheet.cell(
                row=row,
                column=2,
                value=self.analysis.unique_values[column],
            )

            worksheet.cell(
                row=row,
                column=3,
                value=(
                    str(most_frequent)
                    if most_frequent is not None
                    else "N/A"
                ),
            )

            worksheet.cell(
                row=row,
                column=4,
                value=frequency,
            )

            percentage_cell = worksheet.cell(
                row=row,
                column=5,
                value=percentage,
            )

            percentage_cell.number_format = "0.00%"

            worksheet.cell(
                row=row,
                column=6,
                value=self.analysis.null_values[column],
            )
            
    def _build_temporal_dashboard(
        self,
        dashboard,
        data_sheet,
    ) -> None:
        date_column = self.analysis.datetime_columns[0]

        series = self.df[date_column].dropna()

        if series.empty:
            return

        start_date = series.min()
        end_date = series.max()

        duration_days = (
            end_date - start_date
        ).days

        kpis = [
            (
                "Registros",
                self.analysis.rows,
            ),
            (
                "Data inicial",
                start_date,
            ),
            (
                "Data final",
                end_date,
            ),
            (
                "Duração (dias)",
                duration_days,
            ),
        ]

        self._write_kpis(
            dashboard,
            kpis,
        )

        metric = self.analysis.primary_metric

        if metric:
            self._build_temporal_metric_chart(
                dashboard,
                data_sheet,
                date_column,
                metric,
            )

            self._build_temporal_metric_summary(
                dashboard,
                date_column,
                metric,
            )

        else:
            self._build_event_count_chart(
                dashboard,
                data_sheet,
                date_column,
            )

            self._build_temporal_event_summary(
                dashboard,
                date_column,
            )
            
    def _build_temporal_metric_chart(
        self,
        dashboard,
        data_sheet,
        date_column,
        metric,
    ) -> None:
        grouped = (
            self.df[
                [date_column, metric]
            ]
            .dropna()
            .groupby(
                date_column,
                as_index=False,
            )[metric]
            .sum()
            .sort_values(date_column)
        )

        if grouped.empty:
            return

        data_sheet["A1"] = date_column
        data_sheet["B1"] = metric

        for row, (_, record) in enumerate(
            grouped.iterrows(),
            start=2,
        ):
            date_cell = data_sheet.cell(
                row=row,
                column=1,
                value=record[date_column],
            )

            date_cell.number_format = "DD/MM/YYYY"

            data_sheet.cell(
                row=row,
                column=2,
                value=float(record[metric]),
            )

        chart = LineChart()

        chart.title = (
            f"{metric.title()} ao longo do tempo"
        )

        chart.x_axis.title = (
            date_column.title()
        )

        chart.y_axis.title = (
            metric.title()
        )

        chart.height = 10
        chart.width = 20

        data = Reference(
            data_sheet,
            min_col=2,
            min_row=1,
            max_row=len(grouped) + 1,
        )

        categories = Reference(
            data_sheet,
            min_col=1,
            min_row=2,
            max_row=len(grouped) + 1,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "A9",
        )
        
    
    def _build_event_count_chart(
        self,
        dashboard,
        data_sheet,
        date_column,
    ) -> None:
        grouped = (
            self.df[[date_column]]
            .dropna()
            .groupby(date_column)
            .size()
            .reset_index(
                name="eventos"
            )
            .sort_values(date_column)
        )

        if grouped.empty:
            return

        data_sheet["D1"] = date_column
        data_sheet["E1"] = "eventos"

        for row, (_, record) in enumerate(
            grouped.iterrows(),
            start=2,
        ):
            date_cell = data_sheet.cell(
                row=row,
                column=4,
                value=record[date_column],
            )

            date_cell.number_format = "DD/MM/YYYY"

            data_sheet.cell(
                row=row,
                column=5,
                value=int(record["eventos"]),
            )

        chart = LineChart()

        chart.title = "Eventos ao longo do tempo"
        chart.x_axis.title = date_column.title()
        chart.y_axis.title = "Eventos"

        chart.height = 10
        chart.width = 20

        data = Reference(
            data_sheet,
            min_col=5,
            min_row=1,
            max_row=len(grouped) + 1,
        )

        categories = Reference(
            data_sheet,
            min_col=4,
            min_row=2,
            max_row=len(grouped) + 1,
        )

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(categories)

        dashboard.add_chart(
            chart,
            "A9",
        )
        
        
    def _build_temporal_metric_summary(
        self,
        worksheet,
        date_column,
        metric,
    ) -> None:
        grouped = (
            self.df[
                [date_column, metric]
            ]
            .dropna()
            .groupby(
                date_column,
                as_index=False,
            )[metric]
            .sum()
        )

        if grouped.empty:
            return

        peak_row = grouped.loc[
            grouped[metric].idxmax()
        ]

        average = float(
            grouped[metric].mean()
        )

        start_row = 28

        worksheet.cell(
            row=start_row,
            column=1,
            value="Resumo temporal",
        ).font = Font(
            bold=True,
            size=14,
        )

        values = [
            (
                "Métrica",
                metric,
            ),
            (
                "Média por período",
                average,
            ),
            (
                "Maior valor",
                float(peak_row[metric]),
            ),
            (
                "Data do pico",
                peak_row[date_column],
            ),
        ]

        for offset, (label, value) in enumerate(
            values,
            start=2,
        ):
            worksheet.cell(
                row=start_row + offset,
                column=1,
                value=label,
            ).font = Font(bold=True)

            cell = worksheet.cell(
                row=start_row + offset,
                column=2,
                value=value,
            )

            if label == "Data do pico":
                cell.number_format = "DD/MM/YYYY"

            elif isinstance(value, float):
                cell.number_format = "#,##0.00"
                
    
    def _build_temporal_event_summary(
        self,
        worksheet,
        date_column,
    ) -> None:
        grouped = (
            self.df[[date_column]]
            .dropna()
            .groupby(date_column)
            .size()
            .reset_index(
                name="eventos"
            )
        )

        if grouped.empty:
            return

        peak_row = grouped.loc[
            grouped["eventos"].idxmax()
        ]

        average = float(
            grouped["eventos"].mean()
        )

        start_row = 28

        worksheet.cell(
            row=start_row,
            column=1,
            value="Resumo temporal",
        ).font = Font(
            bold=True,
            size=14,
        )

        values = [
            (
                "Média de eventos",
                average,
            ),
            (
                "Maior volume",
                int(peak_row["eventos"]),
            ),
            (
                "Data do pico",
                peak_row[date_column],
            ),
        ]

        for offset, (label, value) in enumerate(
            values,
            start=2,
        ):
            worksheet.cell(
                row=start_row + offset,
                column=1,
                value=label,
            ).font = Font(bold=True)

            cell = worksheet.cell(
                row=start_row + offset,
                column=2,
                value=value,
            )

            if label == "Data do pico":
                cell.number_format = "DD/MM/YYYY"

            elif isinstance(value, float):
                cell.number_format = "#,##0.00"
